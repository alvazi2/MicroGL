import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from database import Database
import json

# Load constants from JSON file
with open('constants.json') as f:
    constants = json.load(f)

DB_PATH = constants['gldbFilePath']
YEAR = '2024'
EXCEL_PATH = constants['excelWriter']['excelPath']
SHEET_NAME = constants['excelWriter']['sheetName']
TABLE_NAME = constants['excelWriter']['tableName']

def write_data_to_excel(db_path, year, excel_path, sheet_name, table_name):
    # Initialize the database
    db = Database(db_path)
    
    # Query the data for the specified year
    query = f"SELECT * FROM gl_items WHERE posting_year = ?"
    df = pd.read_sql_query(query, db.connection, params=(year,))
    
    # Close the database connection
    db.close()
    
    # Load the existing workbook
    book = load_workbook(excel_path)
    
    # Write the dataframe data to the specified sheet
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Reload the workbook to add the table
    book = load_workbook(excel_path)
    sheet = book[sheet_name]
    
    # Remove the existing table if it exists
    existing_tables = [tbl for tbl in sheet.tables.values() if tbl.displayName == table_name]
    for tbl in existing_tables:
        del sheet.tables[tbl.displayName]
    
    # Define the table range
    (max_row, max_col) = df.shape
    table_range = f"A1:{chr(65 + max_col - 1)}{max_row + 1}"
    
    # Create a table
    table = Table(displayName=table_name, ref=table_range)
    
    # Add a default style with striped rows
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    
    # Add the table to the sheet
    sheet.add_table(table)
    
    # Set the column width for better visibility
    for i, col in enumerate(df.columns):
        max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
        sheet.column_dimensions[chr(65 + i)].width = max_len
    
    # Save the workbook
    book.save(excel_path)

write_data_to_excel(DB_PATH, YEAR, EXCEL_PATH, SHEET_NAME, TABLE_NAME)