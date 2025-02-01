import sqlite3
import pandas as pd
from decimal import Decimal
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

DB_PATH = 'alvaziGL_Data/alvaziGL.db'
YEAR = '2024'
EXCEL_PATH = 'alvaziGL_Data/alvaziGL.xlsx'
SHEET_NAME = '2024'
TABLE_NAME = 'Transactions2024'

# Adapter to convert Decimal to integer (scaled by 100)
def adapt_decimal(d):
    return int(d * 100)

# Converter to convert integer back to Decimal (scaled by 100)
def convert_decimal(i):
    return Decimal(i.decode('utf-8')) / 100

def write_data_to_excel(db_path, year, excel_path, sheet_name, table_name):
    # Connect to the SQLite database
    conn = sqlite3.connect(db_path, detect_types=sqlite3.PARSE_DECLTYPES)
    
    # Query the data for the specified year
    query = f"SELECT * FROM gl_items WHERE posting_year = '{year}'"
    df = pd.read_sql_query(query, conn)
    
    # Close the database connection
    conn.close()
    
    # Load the existing workbook
    book = load_workbook(excel_path)
    
    # Write the dataframe data to the specified sheet
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Reload the workbook to add the table
    book = load_workbook(excel_path)
    sheet = book[sheet_name]
    
    # Remove the existing table if it exists
    existing_tables = [tbl for tbl in sheet.tables.values() if tbl.displayName == table_name]
    for tbl in existing_tables:
        del sheet.tables[tbl.displayName]
    
    # Define the table range
    (max_row, max_col) = df.shape
    table_range = f"A1:{chr(65 + max_col - 1)}{max_row + 1}"
    
    # Create a table
    table = Table(displayName=table_name, ref=table_range)
    
    # Add a default style with striped rows
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    
    # Add the table to the sheet
    sheet.add_table(table)
    
    # Set the column width for better visibility
    for i, col in enumerate(df.columns):
        max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
        sheet.column_dimensions[chr(65 + i)].width = max_len
    
    # Save the workbook
    book.save(excel_path)

# Register the adapter and converter for transaction amount
sqlite3.register_adapter(Decimal, adapt_decimal)
sqlite3.register_converter("DECIMAL", convert_decimal)
    
write_data_to_excel(DB_PATH, YEAR, EXCEL_PATH, SHEET_NAME, TABLE_NAME)